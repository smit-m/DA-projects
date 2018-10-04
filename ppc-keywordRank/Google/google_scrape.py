# -*- coding: utf-8 -*-
"""
Created on Wed Jul 25 11:38:07 2018

@author: smehta
"""

from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from fake_useragent import UserAgent
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from datetime import datetime
from openpyxl import load_workbook


################################

client = 'ABC'  #Enter the name of the client

################################

#Read the Adwords-exported workbook
wb = load_workbook(filename='SQR_for_Crawler_xlsx.xlsx', read_only=False)
ws = wb['Sheet0'] #SheetName

row_count = ws.max_row

#Initialize lists that will hold terms
st = [] #Searchterm: Col A
mt = [] #MatchType: Col B
cn = [] #CampaignName: Col D
ag = [] #Ad Group: Col E
kw = [] #Keyword: Col L

for i in range(2,row_count+1): #Starting at 2 since the first row is the header
    st.append(ws['A'+str(i)].value)
    mt.append(ws['B'+str(i)].value)
    cn.append(ws['D'+str(i)].value.replace('|', '-'))
    ag.append(ws['E'+str(i)].value)
    if ws['L'+str(i)].value == '--':
        kw.append(ws['A'+str(i)].value)
    else:
        kw.append(ws['L'+str(i)].value)




#make incognito browser
ua=UserAgent()
dcap = dict(DesiredCapabilities.PHANTOMJS)
dcap["phantomjs.page.settings.userAgent"] = (ua.random)
service_args=['--ssl-protocol=any','--ignore-ssl-errors=true']

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")

driver = webdriver.Chrome('chromedriver',desired_capabilities=dcap,service_args=service_args, options=chrome_options)



#############################

#### Opening the search page directly using the link ####

##### Searching for keywords #####

a = open("keywords_python_v4.txt", 'a')

for i in range(0,row_count-1):
    
    #for keyword in keywords:
        
        driver.get('https://www.google.com/search?q=' + st[i].strip().replace(' ', '+'))
        time.sleep(1.5)       
        counter = 1
            
        els=driver.find_elements_by_class_name('UdQCqe')
        try:
            if els == []:
                a.write(str(datetime.now()) + '|' + client + '|' + st[i] + '|' + mt[i] + '|' + cn[i] + '|' + ag[i] + '|' + kw[i] + '|' + '999' + '|' + 'BadSearch' + '\n')
            else:
                for el in els:
                    a.write(str(datetime.now()) + '|' + client + '|' + st[i] + '|' + mt[i] + '|' + cn[i] + '|' + ag[i] + '|' + kw[i] + '|' + str(counter) + '|' + el.text + '\n')
                    counter = counter + 1
        except:
            pass

            
a.close()


#############################

'''
####### Discarded #######

def shufflewordorder1(x):
    y = x.split()
    l1 = []
    for i in range(len(y)+1):
        random.shuffle(y)
        z = ' '.join(y)
        l1.append(z)
    
    return l1


def transformkeyword1(x):
    if '+' in x:
        temp = x.replace('+', '').strip()
        y = shufflewordorder1(temp)
    elif '[' in x:
        y = x.replace('[', '').replace(']', '').strip()
    elif '"' in x:
        y = 'buy ' + x.replace('"', '').strip() + ' now'
        
    return y
    
'''
